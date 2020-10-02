# 0. Read args from sys.argv
# 1. Create a table
# 2. Create a function to get a int (size of the table) and return list of lists which makes the table and the border
# 3. Create a font object
# 4. Put borders, go through cells in a cycle and put values into them
# 5. Apply font to 1st column of table and 1st row of the table


def multi_table(size):
	"""
	:param size: size of the table to be generated
	:return: list of values to build multiplication table
	"""
	return [a * b for a in range(1, size + 1) for b in range(1, size + 1)]


import sys, openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

try:
	table_size = int(sys.argv[1])  # 0 element is script name
	print(f"Table size is {table_size}")
except Exception as e:
	print("Expected argument was not passed to script, default table size of 10 will be used")
	table_size = 10

wb = openpyxl.Workbook()  # create new WB
border_font = Font(bold=True)  # create font object
active_sheet = wb.active  # get active sheet, well there are only one yet :)
print(type(active_sheet))
active_sheet.title = "Multiplication_Table"  # change title

for i in range(2, table_size + 2):  # we create table borders and apply font to these cells
	# Setting horizontal border
	active_sheet.cell(row=1, column=i).value = i - 1
	active_sheet.cell(row=1, column=i).font = border_font
	# Setting vertical border
	active_sheet.cell(row=i, column=1).value = i - 1
	active_sheet.cell(row=i, column=1).font = border_font

m_table_data = list(reversed(multi_table(table_size)))  # we reverse list since we use .pop() and we need to pop from the start
m_table = active_sheet[f"{get_column_letter(2)}2":f"{get_column_letter(table_size + 1)}{table_size + 1}"]  # create a slice that will be used for table

for cell in [c for row in m_table for c in row]:  # row makes m_table, c makes row
	cell.value = m_table_data.pop()
try:
	wb.save("Multiplication_workbook.xlsx")
except PermissionError:
	print("Please close the workbook before trying to overwrite it with script or check access permissions")
	sys.exit("Access denied")